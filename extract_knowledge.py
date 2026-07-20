"""
extract_knowledge.py (patched)
================================
เพิ่ม: extract_md_business_logic() — อ่านไฟล์ .md ที่เป็น Business Logic โดยตรง
แทนที่จะอ่านเฉพาะ JSON/Excel/Python
"""
import json
import os
import re
from pathlib import Path

SOURCE_DIR    = Path("source_files")
KNOWLEDGE_DIR = Path("knowledge")
SOURCE_DIR.mkdir(exist_ok=True)
KNOWLEDGE_DIR.mkdir(exist_ok=True)


def clean_text(text: str) -> str:
    cleaned = re.sub(r'[^\u0E00-\u0E7Fa-zA-Z0-9\s\-_:/\.,\(\)\[\]#*`%=+!@\'"<>{}|→✅⚠️📊🖥️📂▼]', '', text)
    cleaned = re.sub(r'\s{3,}', '\n', cleaned)
    return cleaned.strip()


# ── ใหม่: อ่านไฟล์ Business Logic .md โดยตรง ──────────────────────────────
def extract_md_business_logic(md_path: Path) -> str:
    """
    อ่านไฟล์ .md ที่เป็น Business Logic แล้ว clean และ copy ไปยัง knowledge/
    ไม่ต้องแปลง — เพราะเป็น text-only แล้ว (หลังจากทำ Step 4 ไปแล้ว)
    """
    content = md_path.read_text(encoding="utf-8")
    return clean_text(content)


def extract_grafana_json(json_path: Path) -> str:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    lines = [f"# สรุปจาก Grafana JSON: {json_path.stem}\n"]

    variables = data.get("templating", {}).get("list", [])
    if variables:
        lines.append("## Dropdown (Variables)")
        for v in variables:
            name        = clean_text(v.get("name", "ไม่ทราบชื่อ"))
            v_type      = v.get("type", "")
            multi       = v.get("multi", False)
            select_type = "Multiple Select" if multi else "Single Select"
            lines.append(f"- **{name}** ({v_type}) → {select_type}")
        lines.append("")

    panels = data.get("panels", [])
    if panels:
        lines.append("## Panel ทั้งหมด")
        for p in panels:
            title = clean_text(p.get("title", "ไม่มีชื่อ"))
            ptype = p.get("type", "unknown")
            lines.append(f"\n### Panel: {title}")
            lines.append(f"- ประเภท Panel: `{ptype}`")

            targets = p.get("targets", [])
            for t in targets:
                sql = t.get("rawSql", "")
                if sql:
                    sql_clean = clean_text(sql)
                    lines.append(
                        f"- Query (ย่อ): `{sql_clean[:150]}...`"
                        if len(sql_clean) > 150
                        else f"- Query: `{sql_clean}`"
                    )

            field_config = p.get("fieldConfig", {}).get("defaults", {})
            thresholds = field_config.get("thresholds", {}).get("steps", [])
            if thresholds:
                lines.append("- Threshold สี:")
                for step in thresholds:
                    lines.append(f"  - ค่า {step.get('value', 0)} → สี {step.get('color', 'N/A')}")

            overrides = p.get("fieldConfig", {}).get("overrides", [])
            for ov in overrides:
                matcher_val = ov.get("matcher", {}).get("options", "")
                for prop in ov.get("properties", []):
                    if prop.get("id") == "color":
                        color_val = prop.get("value", {})
                        lines.append(f"- Field '{matcher_val}' → สี: {color_val.get('fixedColor', color_val.get('mode',''))}")

    return "\n".join(lines)


def extract_python_sql(py_path: Path) -> str:
    content   = py_path.read_text(encoding="utf-8", errors="ignore")
    lines_out = [f"# สรุปจาก Python Script: {py_path.name}\n"]
    lines     = content.split("\n")
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("def "):
            func = stripped.split("(")[0].replace("def ", "")
            lines_out.append(f"\n## Function: `{func}()`")
        if stripped.startswith("#") and len(stripped) > 3:
            lines_out.append(f"- หมายเหตุ: {clean_text(stripped.lstrip('#').strip())}")
        if any(kw in stripped.upper() for kw in ["SELECT", "FROM", "WHERE", "JOIN"]):
            lines_out.append(f"- SQL: `{clean_text(stripped[:120])}`")
    return "\n".join(lines_out)


def extract_excel_summary(xlsx_path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return f"# {xlsx_path.name}\n⚠️ ต้องติดตั้ง openpyxl ก่อน: pip install openpyxl"

    wb    = load_workbook(xlsx_path, read_only=True, data_only=True, keep_vba=True)
    lines = [f"# สรุปจาก Excel: {xlsx_path.name}\n"]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"\n## Sheet: {clean_text(sheet_name)}")
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row:
            headers = [clean_text(str(h)) for h in header_row if h is not None]
            lines.append(f"- คอลัมน์: {', '.join(headers)}")
        lines.append("- ตัวอย่างข้อมูล:")
        row_count = 0
        for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
            row_data = [clean_text(str(c)) for c in row if c is not None and str(c).strip()]
            if row_data:
                lines.append(f"  - {', '.join(row_data)}")
                row_count += 1
        if row_count == 0:
            lines.append("  - (ไม่มีข้อมูลในชีตนี้)")
    return "\n".join(lines)


def run_extraction():
    print("=== เริ่ม Extract Knowledge ===\n")

    # ── ใหม่: อ่านไฟล์ Business Logic .md จาก source_files/ ──────────────
    print("[Step 0] Business Logic .md files")
    for md_file in SOURCE_DIR.glob("*.md"):
        print(f"  [Extract] Business Logic MD: {md_file.name}")
        summary  = extract_md_business_logic(md_file)
        # ตั้งชื่อ output ให้ตรงกับที่ KEYWORD_MAP ใน main.py อ้างถึง
        # เช่น d1_qc_member_workload.md → knowledge/d1_qc_member_workload.md
        out_name = md_file.name  # ใช้ชื่อเดิม (ไม่เติม auto_)
        out_path = KNOWLEDGE_DIR / out_name
        out_path.write_text(summary, encoding="utf-8")
        print(f"    → สร้าง: {out_path} ({len(summary)} chars)")
    print()

    # 1. Grafana JSON
    print("[Step 1] Grafana JSON files")
    for json_file in SOURCE_DIR.glob("*.json"):
        print(f"  [Extract] Grafana JSON: {json_file.name}")
        summary  = extract_grafana_json(json_file)
        out_path = KNOWLEDGE_DIR / f"auto_{json_file.stem}.md"
        out_path.write_text(summary, encoding="utf-8")
        print(f"    → สร้าง: {out_path}")
    print()

    # 2. Python Scripts
    print("[Step 2] Python scripts")
    for py_file in SOURCE_DIR.glob("*.py"):
        print(f"  [Extract] Python: {py_file.name}")
        summary  = extract_python_sql(py_file)
        out_path = KNOWLEDGE_DIR / f"auto_{py_file.stem}.md"
        out_path.write_text(summary, encoding="utf-8")
        print(f"    → สร้าง: {out_path}")
    print()

    # 3. Excel Files
    print("[Step 3] Excel files")
    excel_files = list(SOURCE_DIR.glob("*.xlsx")) + list(SOURCE_DIR.glob("*.xlsm"))
    for xlsx_file in excel_files:
        print(f"  [Extract] Excel ({xlsx_file.suffix}): {xlsx_file.name}")
        summary  = extract_excel_summary(xlsx_file)
        out_path = KNOWLEDGE_DIR / f"auto_{xlsx_file.stem}.md"
        out_path.write_text(summary, encoding="utf-8")
        print(f"    → สร้าง: {out_path}")
    print()

    print("=== เสร็จสิ้น ===")
    all_md = list(KNOWLEDGE_DIR.glob("*.md"))
    print(f"ไฟล์ .md ทั้งหมดใน knowledge/: {len(all_md)} ไฟล์")
    for f in sorted(all_md):
        size = len(f.read_text(encoding="utf-8"))
        print(f"  - {f.name} ({size:,} chars)")


if __name__ == "__main__":
    run_extraction()